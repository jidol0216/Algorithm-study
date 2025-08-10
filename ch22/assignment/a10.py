from openpyxl import load_workbook


wb = load_workbook(r"C:/rokey/python/ch22/assignment/data.xlsx")

ws = wb.worksheets[0]

ws["A1"] = "Name"
ws["B1"] = "Age"
ws["C1"] = "City"

ws.append(["Alice", 30, "Seoul"])
ws.append(["Bob", 25, "Busan"])

wb.save(r"C:/rokey/python/ch22/assignment/data.xlsx")

print("data.xlsx 첫 번째 시트에 데이터가 추가되었습니다.")
